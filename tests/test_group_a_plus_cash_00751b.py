from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from group_a_plus.portfolio.cash_00751b import (
    _metrics,
    read_group_a_plus_plus_holdings,
    workbook_snapshot_date,
    workbook_snapshot_is_stale,
)


def _write_workbook(tmp_path: Path, *, trailing_group: bool) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Account")
    ws.cell(1, 2, "Group A++")
    if trailing_group:
        ws.cell(1, 4, "Group B (other)")
    ws.cell(2, 2, "0050")
    ws.cell(2, 3, "00751B")
    ws.cell(3, 2, 100)
    ws.cell(3, 3, 50)
    if trailing_group:
        ws.cell(2, 4, "0056")
        ws.cell(3, 4, 9999)
    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    return path


def test_read_holdings_stops_at_next_group_header(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, trailing_group=True)
    holdings = read_group_a_plus_plus_holdings(path)
    assert holdings == {"0050.TW": 100.0, "00751B.TWO": 50.0}
    assert "0056.TW" not in holdings


def test_read_holdings_falls_back_to_max_column_without_next_group(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, trailing_group=False)
    holdings = read_group_a_plus_plus_holdings(path)
    assert holdings == {"0050.TW": 100.0, "00751B.TWO": 50.0}


def test_read_holdings_raises_without_group_a_plus_plus_header(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Account")
    path = tmp_path / "no_header.xlsx"
    wb.save(path)
    with pytest.raises(RuntimeError, match="Group A\\+\\+ header not found"):
        read_group_a_plus_plus_holdings(path)


def test_metrics_on_constant_series_has_zero_return_and_sharpe() -> None:
    idx = pd.date_range("2025-01-01", periods=30, freq="D")
    values = pd.Series([100.0] * 30, index=idx)
    metrics = _metrics(values)
    assert metrics["total_return"] == pytest.approx(0.0)
    assert metrics["sharpe_ratio"] == pytest.approx(0.0)
    assert metrics["max_drawdown"] == pytest.approx(0.0)
    assert metrics["worst_20d_return"] == pytest.approx(0.0)


def test_metrics_total_return_matches_manual_calculation() -> None:
    idx = pd.date_range("2025-01-01", periods=3, freq="D")
    values = pd.Series([100.0, 110.0, 121.0], index=idx)
    metrics = _metrics(values)
    assert metrics["total_return"] == pytest.approx(0.21)
    assert metrics["start_value"] == pytest.approx(100.0)
    assert metrics["final_value"] == pytest.approx(121.0)


def test_metrics_max_drawdown_is_negative_after_a_drop() -> None:
    idx = pd.date_range("2025-01-01", periods=3, freq="D")
    values = pd.Series([100.0, 50.0, 60.0], index=idx)
    metrics = _metrics(values)
    assert metrics["max_drawdown"] == pytest.approx(-0.5)


def test_workbook_snapshot_date_parses_yyyymmdd_from_filename() -> None:
    assert workbook_snapshot_date(Path("taiwan_stock_20260619.xlsx")) == date(2026, 6, 19)


def test_workbook_snapshot_date_is_none_without_a_date() -> None:
    assert workbook_snapshot_date(Path("taiwan_stock_latest.xlsx")) is None


def test_workbook_snapshot_is_stale_past_max_age() -> None:
    path = Path("taiwan_stock_20260101.xlsx")
    assert workbook_snapshot_is_stale(path, today=date(2026, 3, 1), max_age_days=30) is True


def test_workbook_snapshot_is_not_stale_within_max_age() -> None:
    path = Path("taiwan_stock_20260101.xlsx")
    assert workbook_snapshot_is_stale(path, today=date(2026, 1, 15), max_age_days=30) is False


def test_workbook_snapshot_is_stale_when_date_unparseable() -> None:
    path = Path("taiwan_stock_latest.xlsx")
    assert workbook_snapshot_is_stale(path, today=date(2026, 3, 1)) is True
