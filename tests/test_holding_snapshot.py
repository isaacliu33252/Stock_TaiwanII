from __future__ import annotations

import json

import openpyxl
import pytest

from group_a_plus.portfolio.holding_snapshot import (
    holding_snapshot_from_dict,
    load_holding_snapshot_excel,
    load_holding_snapshot_json,
)


def test_holding_snapshot_from_dict_accepts_holdings() -> None:
    snapshot = holding_snapshot_from_dict(
        {
            "account_id": "paper",
            "as_of": "2026-07-27",
            "cash": 300_000,
            "holdings": {"0050.tw": 4000, "00631l.tw": "12000"},
        },
        source="unit",
    )

    assert snapshot.current_shares == {"0050.TW": 4000.0, "00631L.TW": 12000.0}
    assert snapshot.cash == 300_000.0
    assert snapshot.account_id == "paper"
    assert snapshot.as_of == "2026-07-27"
    assert snapshot.source == "unit"


def test_holding_snapshot_from_dict_accepts_current_shares_alias() -> None:
    snapshot = holding_snapshot_from_dict({"cash": 1, "current_shares": {"0050.TW": 2}})
    assert snapshot.current_shares == {"0050.TW": 2.0}
    assert snapshot.cash == 1.0


def test_holding_snapshot_rejects_negative_cash() -> None:
    with pytest.raises(ValueError, match="cash must be non-negative"):
        holding_snapshot_from_dict({"cash": -1, "holdings": {"0050.TW": 1}})


def test_holding_snapshot_rejects_missing_holdings_object() -> None:
    with pytest.raises(ValueError, match="must contain holdings"):
        holding_snapshot_from_dict({"cash": 1})


def test_load_holding_snapshot_json(tmp_path) -> None:
    path = tmp_path / "holdings.json"
    path.write_text(json.dumps({"cash": 100, "holdings": {"0050.TW": 3}}), encoding="utf-8")

    snapshot = load_holding_snapshot_json(path)

    assert snapshot.current_shares == {"0050.TW": 3.0}
    assert snapshot.cash == 100.0
    assert snapshot.source == str(path)


def test_load_holding_snapshot_excel_table_format(tmp_path) -> None:
    path = tmp_path / "holdings_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ticker", "shares"])
    ws.append(["0050", 4000])
    ws.append(["00631L", "12000"])
    ws.append(["00679B", 1000])
    wb.save(path)

    snapshot = load_holding_snapshot_excel(path, cash=300_000, account_id="excel", as_of="2026-07-29")

    assert snapshot.current_shares == {"0050.TW": 4000.0, "00631L.TW": 12000.0, "00679B.TWO": 1000.0}
    assert snapshot.cash == 300_000.0
    assert snapshot.account_id == "excel"
    assert snapshot.as_of == "2026-07-29"


def test_load_holding_snapshot_excel_table_format_can_parse_cash_row(tmp_path) -> None:
    path = tmp_path / "holdings_table_cash.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ticker", "shares"])
    ws.append(["0050.TW", 4000])
    ws.append(["cash", 123456])
    wb.save(path)

    snapshot = load_holding_snapshot_excel(path)

    assert snapshot.current_shares == {"0050.TW": 4000.0}
    assert snapshot.cash == 123456.0


def test_load_holding_snapshot_excel_horizontal_taiwan_stock_format(tmp_path) -> None:
    path = tmp_path / "taiwan_stock_like.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["", "Group A++", "", "", "", "Group B"])
    ws.append(["", "元大台灣50 \n0050", "元大台灣50正2\n00631L", "元大美債20年\n00679B", "元大台灣50+2\n0063L", "0056"])
    ws.append(["即時庫存", 1342, 0, 5000, 680, 16673])
    wb.save(path)

    snapshot = load_holding_snapshot_excel(path, cash=10_000)

    assert snapshot.current_shares == {
        "0050.TW": 1342.0,
        "00631L.TW": 680.0,
        "00679B.TWO": 5000.0,
    }
    assert snapshot.cash == 10_000.0


def test_load_holding_snapshot_excel_requires_cash_when_workbook_has_none(tmp_path) -> None:
    path = tmp_path / "holdings_no_cash.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ticker", "shares"])
    ws.append(["0050", 4000])
    wb.save(path)

    with pytest.raises(ValueError, match="pass cash explicitly"):
        load_holding_snapshot_excel(path)
