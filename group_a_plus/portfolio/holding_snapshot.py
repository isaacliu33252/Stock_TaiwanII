"""Load broker-neutral holding snapshots for GroupA+ rebalance planning."""

from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


TPEX_CODES = {"00679B", "00751B"}
CODE_ALIASES = {
    "0063L": "00631L",
}
TICKER_HEADER_ALIASES = {"ticker", "symbol", "code", "stock", "stock_id", "股票", "股票代號", "代號"}
SHARES_HEADER_ALIASES = {"shares", "quantity", "qty", "holding", "holdings", "current_shares", "庫存", "即時庫存", "股數"}
CASH_HEADER_ALIASES = {"cash", "現金", "cash_balance", "available_cash", "銀行餘額"}


@dataclass(frozen=True)
class HoldingSnapshot:
    current_shares: dict[str, float]
    cash: float
    source: str
    account_id: str | None = None
    as_of: str | None = None

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "current_shares": dict(self.current_shares),
            "cash": self.cash,
            "source": self.source,
            "account_id": self.account_id,
            "as_of": self.as_of,
        }


def _clean_number(value: Any, *, field_name: str) -> float:
    try:
        out = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be numeric: {value!r}") from exc
    if not math.isfinite(out):
        raise ValueError(f"{field_name} must be finite: {value!r}")
    return out


def _extract_code(value: Any) -> str | None:
    text = str(value or "").upper()
    matches = re.findall(r"\b\d{4,5}[A-Z]?\b", text)
    if not matches:
        return None
    code = matches[-1]
    return CODE_ALIASES.get(code, code)


def _normalize_ticker(value: Any) -> str:
    text = str(value or "").strip().upper()
    if not text:
        raise ValueError("ticker must not be empty")
    if "." in text:
        return text
    code = _extract_code(text) or text
    suffix = ".TWO" if code in TPEX_CODES else ".TW"
    return f"{code}{suffix}"


def _resolve_holdings(payload: dict[str, Any]) -> dict[str, Any]:
    for key in ("holdings", "current_shares", "positions"):
        value = payload.get(key)
        if isinstance(value, dict):
            return value
    raise ValueError("holding snapshot must contain holdings, current_shares, or positions object")


def holding_snapshot_from_dict(payload: dict[str, Any], *, source: str = "dict") -> HoldingSnapshot:
    if not isinstance(payload, dict):
        raise ValueError("holding snapshot payload must be an object")
    cash = _clean_number(payload.get("cash", 0.0), field_name="cash")
    if cash < 0:
        raise ValueError(f"cash must be non-negative: {cash}")

    raw_holdings = _resolve_holdings(payload)
    current_shares: dict[str, float] = {}
    for ticker, shares_raw in raw_holdings.items():
        ticker_text = _normalize_ticker(ticker)
        shares = _clean_number(shares_raw, field_name=f"holdings[{ticker_text}]")
        if shares < 0:
            raise ValueError(f"holdings[{ticker_text}] must be non-negative: {shares}")
        current_shares[ticker_text] = shares
    if not current_shares:
        raise ValueError("holding snapshot must contain at least one holding")

    return HoldingSnapshot(
        current_shares=current_shares,
        cash=cash,
        source=source,
        account_id=str(payload["account_id"]) if payload.get("account_id") is not None else None,
        as_of=str(payload["as_of"]) if payload.get("as_of") is not None else None,
    )


def load_holding_snapshot_json(path: Path) -> HoldingSnapshot:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return holding_snapshot_from_dict(payload, source=str(path))


def _header_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _load_workbook_values(path: Path, sheet: str | None = None) -> list[list[Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _parse_table_excel_rows(rows: list[list[Any]]) -> tuple[dict[str, float], float | None]:
    for header_idx, row in enumerate(rows[:20]):
        keys = [_header_key(value) for value in row]
        ticker_col = next((idx for idx, key in enumerate(keys) if key in TICKER_HEADER_ALIASES), None)
        shares_col = next((idx for idx, key in enumerate(keys) if key in SHARES_HEADER_ALIASES), None)
        cash_col = next((idx for idx, key in enumerate(keys) if key in CASH_HEADER_ALIASES), None)
        if ticker_col is None or shares_col is None:
            continue
        holdings: dict[str, float] = {}
        cash: float | None = None
        for row_idx, data_row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            raw_ticker = data_row[ticker_col] if ticker_col < len(data_row) else None
            if raw_ticker is None or str(raw_ticker).strip() == "":
                continue
            ticker_text = str(raw_ticker).strip()
            if ticker_text.lower() == "cash" or ticker_text in {"現金", "銀行餘額"}:
                raw_cash = data_row[cash_col] if cash_col is not None and cash_col < len(data_row) else data_row[shares_col]
                cash = _clean_number(raw_cash, field_name=f"excel row {row_idx} cash")
                continue
            shares_raw = data_row[shares_col] if shares_col < len(data_row) else None
            shares = _clean_number(shares_raw, field_name=f"excel row {row_idx} shares")
            if shares < 0:
                raise ValueError(f"excel row {row_idx} shares must be non-negative: {shares}")
            ticker = _normalize_ticker(ticker_text)
            holdings[ticker] = holdings.get(ticker, 0.0) + shares
        if holdings:
            return holdings, cash
    raise ValueError("no table-style holdings found in Excel workbook")


def _parse_horizontal_excel_rows(rows: list[list[Any]], *, row_label: str = "即時庫存") -> dict[str, float]:
    group_row_idx: int | None = None
    group_start: int | None = None
    group_end: int | None = None
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            if str(value or "").strip() in {"Group A++", "Group A+"}:
                group_row_idx = row_idx
                group_start = col_idx
                break
        if group_start is not None:
            break

    holdings_row_idx: int | None = None
    for row_idx, row in enumerate(rows):
        if any(str(value or "").strip() == row_label for value in row):
            holdings_row_idx = row_idx
            break
    if holdings_row_idx is None:
        raise ValueError(f"Excel workbook row not found: {row_label}")
    if holdings_row_idx <= 0:
        raise ValueError(f"Excel workbook has no header row above {row_label}")

    header_row_idx = (group_row_idx + 1) if group_row_idx is not None else holdings_row_idx - 1
    header_row = rows[header_row_idx]
    holdings_row = rows[holdings_row_idx]
    if group_start is None:
        group_start = 0
        group_end = max(len(header_row), len(holdings_row))
    else:
        group_end = max(len(header_row), len(holdings_row))
        group_row = rows[group_row_idx or 0]
        for col_idx in range(group_start + 1, len(group_row)):
            value = str(group_row[col_idx] or "").strip()
            if value.startswith("Group "):
                group_end = col_idx
                break

    holdings: dict[str, float] = {}
    for col_idx in range(group_start, group_end):
        header = header_row[col_idx] if col_idx < len(header_row) else None
        code = _extract_code(header)
        if code is None:
            continue
        shares_raw = holdings_row[col_idx] if col_idx < len(holdings_row) else 0
        if shares_raw is None or str(shares_raw).strip() == "":
            shares = 0.0
        else:
            shares = _clean_number(shares_raw, field_name=f"excel column {col_idx + 1} shares")
        if shares < 0:
            raise ValueError(f"excel column {col_idx + 1} shares must be non-negative: {shares}")
        ticker = _normalize_ticker(code)
        holdings[ticker] = holdings.get(ticker, 0.0) + shares
    if not holdings:
        raise ValueError("no horizontal holdings found in Excel workbook")
    return holdings


def load_holding_snapshot_excel(
    path: Path,
    *,
    sheet: str | None = None,
    cash: float | None = None,
    row_label: str = "即時庫存",
    account_id: str | None = None,
    as_of: str | None = None,
) -> HoldingSnapshot:
    rows = _load_workbook_values(path, sheet)
    parsed_cash: float | None = None
    try:
        holdings, parsed_cash = _parse_table_excel_rows(rows)
    except ValueError:
        holdings = _parse_horizontal_excel_rows(rows, row_label=row_label)
    resolved_cash = _clean_number(cash, field_name="cash") if cash is not None else parsed_cash
    if resolved_cash is None:
        raise ValueError("Excel holdings snapshot has no cash value; pass cash explicitly")
    if resolved_cash < 0:
        raise ValueError(f"cash must be non-negative: {resolved_cash}")
    return HoldingSnapshot(
        current_shares=holdings,
        cash=resolved_cash,
        source=str(path),
        account_id=account_id,
        as_of=as_of,
    )
