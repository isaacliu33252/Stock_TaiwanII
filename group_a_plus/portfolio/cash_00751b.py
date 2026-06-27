"""Evaluate whether 00751B is better than cash for current GroupA++ holdings."""

from __future__ import annotations

import argparse
import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb
import openpyxl
import pandas as pd

from backtest_group_a_plus_switch_policy import DB_PATH
from group_a_plus.paths import PROJECT_ROOT


DEFAULT_WORKBOOK = PROJECT_ROOT / "taiwan_stock_20260619.xlsx"

HEADER_TO_TICKER = {
    "0050": "0050.TW",
    "00631L": "00631L.TW",
    "00632R": "00632R.TW",
    "00679B": "00679B.TWO",
    "00751B": "00751B.TWO",
    "0056": "0056.TW",
    "00646": "00646.TW",
    "00713": "00713.TW",
    "00878": "00878.TW",
}


def _extract_code(value: Any) -> str | None:
    text = str(value or "").upper()
    for code in HEADER_TO_TICKER:
        if code in text:
            return code
    return None


def read_group_a_plus_plus_holdings(workbook_path: Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    holdings: dict[str, float] = {}
    group_start = None
    group_end = None
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(1, col).value or "")
        if "Group A++" in value:
            group_start = col
            break
    if group_start is None:
        raise RuntimeError("Group A++ header not found in row 1")
    for col in range(group_start + 1, ws.max_column + 2):
        value = str(ws.cell(1, col).value or "")
        if col > group_start and value.strip().startswith("Group "):
            group_end = col - 1
            break
    if group_end is None:
        group_end = ws.max_column

    for col in range(group_start, group_end + 1):
        code = _extract_code(ws.cell(2, col).value)
        if not code:
            continue
        shares = ws.cell(3, col).value or 0
        try:
            shares_float = float(shares)
        except (TypeError, ValueError):
            shares_float = 0.0
        holdings[HEADER_TO_TICKER[code]] = shares_float
    return holdings


def load_prices(db_path: Path, tickers: list[str], start: str, end: str) -> pd.DataFrame:
    placeholders = ", ".join(["?"] * len(tickers))
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        rows = con.execute(
            f"""
            SELECT dt, ticker, close
            FROM ohlcv
            WHERE ticker IN ({placeholders}) AND dt BETWEEN ? AND ?
            ORDER BY dt, ticker
            """,
            [*tickers, start, end],
        ).fetchdf()
    finally:
        con.close()
    if rows.empty:
        raise RuntimeError(f"No prices for {tickers} between {start} and {end}")
    frame = rows.pivot(index="dt", columns="ticker", values="close").sort_index()
    frame.index = pd.to_datetime(frame.index)
    return frame.dropna(how="all")


def _metrics(values: pd.Series) -> dict[str, float]:
    returns = values.pct_change().dropna()
    years = max((values.index[-1] - values.index[0]).days / 365.25, 1e-9)
    total_return = float(values.iloc[-1] / values.iloc[0] - 1.0)
    annual_return = float((values.iloc[-1] / values.iloc[0]) ** (1.0 / years) - 1.0)
    volatility = float(returns.std() * math.sqrt(252)) if len(returns) > 1 else 0.0
    sharpe = float((returns.mean() / returns.std()) * math.sqrt(252)) if len(returns) > 1 and returns.std() > 0 else 0.0
    max_drawdown = float((values / values.cummax() - 1.0).min())
    worst_20d_return = float(values.pct_change(20).dropna().min()) if len(values) > 20 else 0.0
    return {
        "start_value": float(values.iloc[0]),
        "final_value": float(values.iloc[-1]),
        "total_return": total_return,
        "annual_return": annual_return,
        "volatility": volatility,
        "sharpe_ratio": sharpe,
        "max_drawdown": max_drawdown,
        "worst_20d_return": worst_20d_return,
    }


def evaluate(workbook: Path, db: Path, start: str, end: str) -> tuple[dict[str, Any], pd.DataFrame]:
    holdings = read_group_a_plus_plus_holdings(workbook)
    positive = {ticker: shares for ticker, shares in holdings.items() if shares > 0}
    if "00751B.TWO" not in positive:
        raise RuntimeError("Group A++ has no positive 00751B holding to compare")
    prices = load_prices(db, list(positive), start, end).dropna(subset=list(positive))
    actual_values = pd.DataFrame({ticker: prices[ticker] * shares for ticker, shares in positive.items()}, index=prices.index)
    actual_curve = actual_values.sum(axis=1)
    cash_replacement_value = float(actual_values["00751B.TWO"].iloc[0])
    cash_values = actual_values.drop(columns=["00751B.TWO"]).sum(axis=1) + cash_replacement_value
    bond_curve = actual_values["00751B.TWO"]
    cash_curve = pd.Series(cash_replacement_value, index=prices.index, dtype=float)
    curves = pd.DataFrame(
        {
            "group_a_plus_plus_with_00751b": actual_curve,
            "group_a_plus_plus_00751b_as_cash": cash_values,
            "00751b_only": bond_curve,
            "cash_only_replacing_00751b": cash_curve,
        },
        index=prices.index,
    )
    current_prices = {ticker: float(prices[ticker].iloc[-1]) for ticker in positive}
    current_values = {ticker: float(actual_values[ticker].iloc[-1]) for ticker in positive}
    current_total = sum(current_values.values())
    metrics = {name: _metrics(curves[name]) for name in curves.columns}
    report = {
        "experiment": "group_a_plus_plus_00751b_vs_cash",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(workbook),
        "requested_window": {"start": start, "end": end},
        "actual_window": {"start": str(prices.index[0].date()), "end": str(prices.index[-1].date()), "rows": int(len(prices))},
        "holdings": positive,
        "current_prices": current_prices,
        "current_values": current_values,
        "current_weights": {ticker: value / current_total for ticker, value in current_values.items()},
        "metrics": metrics,
        "comparison": {
            "00751b_minus_cash_final_value": metrics["00751b_only"]["final_value"] - metrics["cash_only_replacing_00751b"]["final_value"],
            "00751b_minus_cash_total_return": metrics["00751b_only"]["total_return"] - metrics["cash_only_replacing_00751b"]["total_return"],
            "portfolio_with_00751b_minus_cash_final_value": metrics["group_a_plus_plus_with_00751b"]["final_value"] - metrics["group_a_plus_plus_00751b_as_cash"]["final_value"],
            "portfolio_with_00751b_minus_cash_total_return": metrics["group_a_plus_plus_with_00751b"]["total_return"] - metrics["group_a_plus_plus_00751b_as_cash"]["total_return"],
        },
        "method_note": "Cash replacement uses the 00751B position value on the first backtest date and assumes zero cash yield. ETF distributions are not added unless already reflected in local close data.",
    }
    return report, curves


def write_workbook(input_path: Path, output_path: Path, report: dict[str, Any], curves: pd.DataFrame) -> None:
    wb = openpyxl.load_workbook(input_path)
    sheet_name = "A++_00751B評估"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Generated", report["generated_at"]])
    ws.append(["Window", f"{report['actual_window']['start']} ~ {report['actual_window']['end']}"])
    ws.append([])
    ws.append(["Scenario", "Start", "Final", "Total Return", "Annual Return", "Volatility", "Sharpe", "MDD", "Worst 20D"])
    for name, metrics in report["metrics"].items():
        ws.append(
            [
                name,
                metrics["start_value"],
                metrics["final_value"],
                metrics["total_return"],
                metrics["annual_return"],
                metrics["volatility"],
                metrics["sharpe_ratio"],
                metrics["max_drawdown"],
                metrics["worst_20d_return"],
            ]
        )
    ws.append([])
    ws.append(["Ticker", "Shares", "Last Price", "Current Value", "Current Weight"])
    for ticker, shares in report["holdings"].items():
        ws.append(
            [
                ticker,
                shares,
                report["current_prices"][ticker],
                report["current_values"][ticker],
                report["current_weights"][ticker],
            ]
        )
    ws.append([])
    ws.append(["Method note", report["method_note"]])
    for col in range(1, 10):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--start", default="2020-01-02")
    parser.add_argument("--end", default="2026-06-18")
    parser.add_argument("--output-prefix", default="results/group_a_plus_plus_00751b_cash_20260619")
    parser.add_argument("--workbook-output", default="taiwan_stock_20260619_groupA++_00751B_eval.xlsx")
    args = parser.parse_args()
    report, curves = evaluate(Path(args.workbook), Path(args.db), args.start, args.end)
    prefix = Path(args.output_prefix)
    prefix.parent.mkdir(parents=True, exist_ok=True)
    json_path = prefix.with_suffix(".json")
    curve_path = prefix.with_name(prefix.name + "_curve.csv")
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    curves.to_csv(curve_path, encoding="utf-8-sig")
    write_workbook(Path(args.workbook), Path(args.workbook_output), report, curves)
    print(f"JSON: {json_path.resolve()}")
    print(f"Curve: {curve_path.resolve()}")
    print(f"Workbook: {Path(args.workbook_output).resolve()}")
    for name, metrics in report["metrics"].items():
        print(
            f"{name}: final={metrics['final_value']:,.0f}, return={metrics['total_return']:.2%}, "
            f"sharpe={metrics['sharpe_ratio']:.3f}, mdd={metrics['max_drawdown']:.2%}"
        )


if __name__ == "__main__":
    main()

