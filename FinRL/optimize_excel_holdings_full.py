#!/usr/bin/env python3
"""Optimize a Taiwan holdings workbook across all current holdings plus 00631L/00632R overlays."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from math import sqrt
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
RESULTS_DIR = PROJECT_ROOT / "results"

BASE_ASSET_CONFIG = [
    {"code": "0050", "ticker": "0050.TW", "name": "元大台灣50"},
    {"code": "0056", "ticker": "0056.TW", "name": "元大高股息"},
    {"code": "00646", "ticker": "00646.TW", "name": "元大S&P500"},
    {"code": "00679B", "ticker": "00679B.TWO", "name": "元大美債20年"},
    {"code": "00713", "ticker": "00713.TW", "name": "元大台灣高息低波"},
    {"code": "00751B", "ticker": "00751B.TWO", "name": "元大AAA至A公司債"},
    {"code": "00878", "ticker": "00878.TW", "name": "國泰永續高股息"},
    {"code": "2884", "ticker": "2884.TW", "name": "玉山金"},
    {"code": "00632R", "ticker": "00632R.TW", "name": "元大台灣50反1"},
    {"code": "00631L", "ticker": "00631L.TW", "name": "元大台灣50正2"},
]

PROFILE_PRESETS = {
    "aggressive": {
        "bounds": {
            "0050.TW": {"dirichlet_alpha": 1.3, "min_weight": 0.08, "max_weight": 0.24},
            "0056.TW": {"dirichlet_alpha": 1.5, "min_weight": 0.08, "max_weight": 0.28},
            "00646.TW": {"dirichlet_alpha": 0.7, "min_weight": 0.02, "max_weight": 0.12},
            "00679B.TWO": {"dirichlet_alpha": 1.2, "min_weight": 0.08, "max_weight": 0.25},
            "00713.TW": {"dirichlet_alpha": 1.0, "min_weight": 0.05, "max_weight": 0.18},
            "00751B.TWO": {"dirichlet_alpha": 0.8, "min_weight": 0.02, "max_weight": 0.10},
            "00878.TW": {"dirichlet_alpha": 1.3, "min_weight": 0.08, "max_weight": 0.20},
            "2884.TW": {"dirichlet_alpha": 0.15, "min_weight": 0.00, "max_weight": 0.04},
            "00632R.TW": {"dirichlet_alpha": 0.06, "min_weight": 0.00, "max_weight": 0.04},
            "00631L.TW": {"dirichlet_alpha": 0.5, "min_weight": 0.02, "max_weight": 0.14},
        },
        "overlay_cap": 0.16,
        "bond_range": {"min": 0.10, "max": 0.35},
        "beta_range": {"min": 0.40, "max": 0.75},
        "score_weights": {"cagr": 1.25, "sharpe": 0.60, "mdd": 0.75, "turnover": 0.10},
    },
    "conservative": {
        "bounds": {
            "0050.TW": {"dirichlet_alpha": 1.4, "min_weight": 0.06, "max_weight": 0.18},
            "0056.TW": {"dirichlet_alpha": 1.7, "min_weight": 0.10, "max_weight": 0.25},
            "00646.TW": {"dirichlet_alpha": 0.8, "min_weight": 0.02, "max_weight": 0.08},
            "00679B.TWO": {"dirichlet_alpha": 1.8, "min_weight": 0.12, "max_weight": 0.30},
            "00713.TW": {"dirichlet_alpha": 1.1, "min_weight": 0.06, "max_weight": 0.15},
            "00751B.TWO": {"dirichlet_alpha": 1.0, "min_weight": 0.04, "max_weight": 0.14},
            "00878.TW": {"dirichlet_alpha": 1.6, "min_weight": 0.10, "max_weight": 0.22},
            "2884.TW": {"dirichlet_alpha": 0.12, "min_weight": 0.00, "max_weight": 0.03},
            "00632R.TW": {"dirichlet_alpha": 0.10, "min_weight": 0.00, "max_weight": 0.05},
            "00631L.TW": {"dirichlet_alpha": 0.2, "min_weight": 0.00, "max_weight": 0.08},
        },
        "overlay_cap": 0.10,
        "bond_range": {"min": 0.20, "max": 0.42},
        "beta_range": {"min": 0.25, "max": 0.60},
        "score_weights": {"cagr": 1.00, "sharpe": 0.85, "mdd": 1.15, "turnover": 0.08},
    },
}

TICKERS = [item["ticker"] for item in BASE_ASSET_CONFIG]
CODE_TO_TICKER = {item["code"]: item["ticker"] for item in BASE_ASSET_CONFIG}
TICKER_TO_CODE = {item["ticker"]: item["code"] for item in BASE_ASSET_CONFIG}
TICKER_TO_NAME = {item["ticker"]: item["name"] for item in BASE_ASSET_CONFIG}
BENCHMARK_TICKER = "0050.TW"


@dataclass
class Candidate:
    score: float
    turnover: float
    beta: float
    weights: np.ndarray
    train_metrics: dict[str, float]
    test_metrics: dict[str, float]


def _safe_int_shares(value: object) -> int:
    if value is None or pd.isna(value):
        return 0
    return int(float(value))


def _resolve_profile(
    profile_name: str,
    excluded_codes: set[str] | None = None,
    min_weight_floor: float = 0.0,
    max_weight_cap: float = 1.0,
    replace_max_weights_with_cap: bool = False,
) -> dict:
    if profile_name not in PROFILE_PRESETS:
        raise ValueError(f"Unsupported profile: {profile_name}")
    excluded_codes = excluded_codes or set()
    preset = PROFILE_PRESETS[profile_name]
    assets = []
    for asset in BASE_ASSET_CONFIG:
        if asset["code"] in excluded_codes:
            continue
        bounds = dict(preset["bounds"][asset["ticker"]])
        bounds["min_weight"] = max(float(bounds["min_weight"]), float(min_weight_floor))
        if replace_max_weights_with_cap:
            bounds["max_weight"] = float(max_weight_cap)
        else:
            bounds["max_weight"] = min(float(bounds["max_weight"]), float(max_weight_cap))
        if bounds["min_weight"] > float(bounds["max_weight"]):
            raise ValueError(
                f"Weight constraints are infeasible for {asset['code']}: "
                f"min={bounds['min_weight']:.2%}, max={bounds['max_weight']:.2%}"
            )
        assets.append({**asset, **bounds})
    return {
        "name": profile_name,
        "assets": assets,
        "excluded_codes": sorted(excluded_codes),
        "min_weight_floor": float(min_weight_floor),
        "max_weight_cap": float(max_weight_cap),
        "replace_max_weights_with_cap": bool(replace_max_weights_with_cap),
        "overlay_cap": float(preset["overlay_cap"]),
        "bond_range": dict(preset["bond_range"]),
        "beta_range": dict(preset["beta_range"]),
        "score_weights": dict(preset["score_weights"]),
    }


def _header_code(cell_value: object) -> str | None:
    if cell_value is None:
        return None
    text = str(cell_value).strip()
    if not text:
        return None
    return text.split("\n")[-1].strip().upper()


def _ensure_main_sheet_columns(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb[wb.sheetnames[0]]
    existing_codes = {}
    for col_idx in range(2, ws.max_column + 1):
        code = _header_code(ws.cell(row=1, column=col_idx).value)
        if code:
            existing_codes[code] = col_idx

    next_col = ws.max_column + 1
    changed = False
    for asset in BASE_ASSET_CONFIG:
        if asset["code"] in existing_codes:
            continue
        ws.cell(row=1, column=next_col, value=f'{asset["name"]}\n{asset["code"]}')
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=next_col, value=0)
        next_col += 1
        changed = True

    if changed:
        wb.save(workbook_path)


def _load_current_holdings(xlsx_path: Path) -> tuple[pd.DataFrame, list[str]]:
    df = pd.read_excel(xlsx_path)
    if df.empty:
        raise ValueError(f"Workbook has no holdings rows: {xlsx_path}")
    codes = [_header_code(col) for col in df.columns[1:]]
    current_row = df.iloc[0]
    rows = []
    for code, col in zip(codes, df.columns[1:]):
        if code not in CODE_TO_TICKER:
            continue
        rows.append(
            {
                "code": code,
                "ticker": CODE_TO_TICKER[code],
                "name": TICKER_TO_NAME[CODE_TO_TICKER[code]],
                "current_shares": _safe_int_shares(current_row[col]),
            }
        )
    holdings = pd.DataFrame(rows)
    if holdings.empty:
        raise ValueError(f"No supported columns found in workbook: {xlsx_path}")
    return holdings, codes


def _download_adjusted_close(
    tickers: list[str],
    start: str,
    end: str,
    *,
    preserve_late_start: bool = False,
) -> pd.DataFrame:
    data = yf.download(tickers, start=start, end=end, progress=False, auto_adjust=True)["Close"]
    if isinstance(data, pd.Series):
        data = data.to_frame(name=tickers[0])
    data = data[tickers].sort_index()
    if preserve_late_start:
        raw = data.copy()
        data = data.ffill()
        for ticker in tickers:
            first_valid = raw[ticker].first_valid_index()
            if first_valid is None:
                raise RuntimeError(f"No price history downloaded for {ticker}")
            data.loc[data.index < first_valid, ticker] = np.nan
        data = data.dropna(how="all")
    else:
        data = data.ffill().dropna()
    if data.empty:
        raise RuntimeError("No price history downloaded")
    return data


def _prepare_price_window(
    prices: pd.DataFrame,
    start: str,
    end: str,
    *,
    allow_late_start_assets: bool,
) -> pd.DataFrame:
    window = prices.loc[pd.Timestamp(start):pd.Timestamp(end)].copy()
    if window.empty:
        raise RuntimeError(f"No prices available for range {start} ~ {end}")
    if allow_late_start_assets:
        benchmark = window[BENCHMARK_TICKER].dropna()
        if benchmark.empty:
            raise RuntimeError(f"No benchmark prices available for {BENCHMARK_TICKER} in range {start} ~ {end}")
        window = window.loc[benchmark.index]
    else:
        window = window.dropna()
    if window.empty:
        raise RuntimeError(f"No aligned prices available for range {start} ~ {end}")
    return window


def _returns_from_prices(prices: pd.DataFrame, *, allow_late_start_assets: bool) -> pd.DataFrame:
    returns = prices.pct_change(fill_method=None)
    return returns.iloc[1:] if allow_late_start_assets else returns.dropna()


def _portfolio_metrics(
    returns: pd.DataFrame,
    weights: np.ndarray,
    *,
    allow_late_start_assets: bool = False,
) -> dict[str, float]:
    matrix = returns.to_numpy(dtype=float)
    if allow_late_start_assets:
        matrix = np.nan_to_num(matrix, nan=0.0)
    elif np.isnan(matrix).any():
        raise RuntimeError("NaN returns encountered without late-start handling")
    portfolio_returns = pd.Series(
        matrix @ np.asarray(weights, dtype=float),
        index=returns.index,
    )
    nav = (1.0 + portfolio_returns).cumprod()
    years = len(portfolio_returns) / 252.0
    cagr = float(nav.iloc[-1] ** (1.0 / years) - 1.0) if years > 0 else 0.0
    vol = float(portfolio_returns.std(ddof=1) * sqrt(252.0))
    sharpe = float(cagr / vol) if vol > 0 else 0.0
    peak = nav.cummax()
    max_drawdown = float((nav / peak - 1.0).min())
    return {
        "cagr": cagr,
        "volatility": vol,
        "sharpe": sharpe,
        "max_drawdown": max_drawdown,
        "final_nav": float(nav.iloc[-1]),
    }


def _estimate_betas(train_returns: pd.DataFrame) -> dict[str, float]:
    benchmark = train_returns[BENCHMARK_TICKER]
    betas = {}
    for ticker in train_returns.columns:
        paired = pd.concat([train_returns[ticker], benchmark], axis=1, join="inner").dropna()
        if len(paired) < 20:
            raise RuntimeError(f"Cannot estimate beta for {ticker} because overlapping history is too short")
        variance = float(np.var(paired.iloc[:, 1].to_numpy(dtype=float)))
        if variance <= 0:
            raise RuntimeError(f"Cannot estimate beta for {ticker} because {BENCHMARK_TICKER} variance is zero")
        cov = float(np.cov(paired.iloc[:, 0].to_numpy(dtype=float), paired.iloc[:, 1].to_numpy(dtype=float))[0, 1])
        betas[ticker] = cov / variance
    return betas


def _is_feasible(
    weights: np.ndarray,
    assets: list[dict],
    betas: dict[str, float],
    min_beta: float,
    max_beta: float,
    overlay_cap: float,
    bond_range: dict[str, float],
) -> tuple[bool, float]:
    for asset in assets:
        weight = float(weights[TICKERS.index(asset["ticker"])])
        if weight < asset["min_weight"] or weight > asset["max_weight"]:
            return False, 0.0

    overlay_total = float(weights[TICKERS.index("00631L.TW")] + weights[TICKERS.index("00632R.TW")])
    if overlay_total > overlay_cap:
        return False, 0.0

    bond_total = float(weights[TICKERS.index("00679B.TWO")] + weights[TICKERS.index("00751B.TWO")])
    if bond_total < bond_range["min"] or bond_total > bond_range["max"]:
        return False, 0.0

    beta = float(sum(weights[i] * betas[TICKERS[i]] for i in range(len(TICKERS))))
    if beta < min_beta or beta > max_beta:
        return False, beta
    return True, beta


def _search_best_candidate(
    train_returns: pd.DataFrame,
    test_returns: pd.DataFrame,
    current_weights: np.ndarray,
    *,
    profile: dict,
    samples: int,
    seed: int,
    min_beta: float,
    max_beta: float,
    allow_late_start_assets: bool = False,
) -> Candidate:
    rng = np.random.default_rng(seed)
    betas = _estimate_betas(train_returns)
    assets = profile["assets"]
    dirichlet_alpha = np.array([asset["dirichlet_alpha"] for asset in assets], dtype=float)
    score_weights = profile["score_weights"]
    best: Candidate | None = None

    for _ in range(samples):
        weights_active = rng.dirichlet(dirichlet_alpha)
        weights = np.zeros(len(TICKERS), dtype=float)
        for idx, asset in enumerate(assets):
            weights[TICKERS.index(asset["ticker"])] = weights_active[idx]
        feasible, beta = _is_feasible(
            weights,
            assets,
            betas,
            min_beta,
            max_beta,
            profile["overlay_cap"],
            profile["bond_range"],
        )
        if not feasible:
            continue

        train_metrics = _portfolio_metrics(
            train_returns,
            weights,
            allow_late_start_assets=allow_late_start_assets,
        )
        test_metrics = _portfolio_metrics(
            test_returns,
            weights,
            allow_late_start_assets=allow_late_start_assets,
        )
        turnover = float(np.abs(weights - current_weights).sum()) / 2.0
        score = (
            score_weights["cagr"] * train_metrics["cagr"]
            + score_weights["sharpe"] * train_metrics["sharpe"]
            - score_weights["mdd"] * abs(train_metrics["max_drawdown"])
            - score_weights["turnover"] * turnover
        )
        candidate = Candidate(
            score=float(score),
            turnover=turnover,
            beta=beta,
            weights=weights.copy(),
            train_metrics=train_metrics,
            test_metrics=test_metrics,
        )
        if best is None or candidate.score > best.score:
            best = candidate

    if best is None:
        raise RuntimeError("No feasible candidate found. Relax constraints or increase samples.")
    return best


def _allocate_integer_shares(
    total_value: float,
    weights: np.ndarray,
    prices: pd.Series,
    *,
    nonzero_tickers: set[str] | None = None,
    eligible_tickers: set[str] | None = None,
) -> pd.Series:
    target_value = total_value * pd.Series(weights, index=prices.index, dtype=float)
    min_shares = pd.Series(0, index=prices.index, dtype=int)
    nonzero_tickers = nonzero_tickers or set()
    eligible_tickers = eligible_tickers or set(prices.index)
    for ticker in nonzero_tickers:
        if ticker not in prices.index:
            continue
        price = float(prices[ticker])
        if price <= 0:
            raise RuntimeError(f"Invalid latest price for {ticker}: {price}")
        min_shares.loc[ticker] = 1

    min_cost = float((min_shares * prices).sum())
    if min_cost > total_value + 1e-9:
        raise RuntimeError("Portfolio value is too small to enforce nonzero target shares")

    residual_budget = float(total_value - min_cost)
    residual_target_value = (target_value - min_shares * prices).clip(lower=0.0)
    if residual_budget > 0 and float(residual_target_value.sum()) > 0:
        scaled_target_value = residual_budget * residual_target_value / float(residual_target_value.sum())
    elif residual_budget > 0 and float(target_value.sum()) > 0:
        scaled_target_value = residual_budget * target_value / float(target_value.sum())
    else:
        scaled_target_value = pd.Series(0.0, index=prices.index)

    exact_shares = scaled_target_value / prices
    extra_shares = np.floor(exact_shares).astype(int)
    shares = min_shares + extra_shares
    residual_cash = float(total_value - float((shares * prices).sum()))
    fractions = (exact_shares - extra_shares)
    candidate_tickers = [
        ticker
        for ticker in fractions.sort_values(ascending=False).index
        if ticker in eligible_tickers and float(target_value[ticker]) > 0.0
    ]
    if not candidate_tickers:
        return pd.Series(shares, index=prices.index, dtype=int)
    min_price = float(prices.loc[candidate_tickers].min())
    while residual_cash >= min_price:
        moved = False
        for ticker in candidate_tickers:
            price = float(prices[ticker])
            if residual_cash + 1e-9 >= price:
                shares.loc[ticker] += 1
                residual_cash -= price
                moved = True
        if not moved:
            break
    return pd.Series(shares, index=prices.index, dtype=int)


def _update_main_sheet(
    workbook_path: Path,
    holdings_by_code: dict[str, int],
    delta_by_code: dict[str, int],
    *,
    target_row_label: str,
    delta_row_label: str,
) -> None:
    wb = load_workbook(workbook_path)
    ws = wb[wb.sheetnames[0]]
    labels = {
        str(ws.cell(row=row_idx, column=1).value).strip(): row_idx
        for row_idx in range(1, ws.max_row + 1)
    }

    def ensure_row(label: str) -> int:
        if label in labels:
            return labels[label]
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=label)
        labels[label] = row_idx
        return row_idx

    target_row = ensure_row(target_row_label)
    delta_row = ensure_row(delta_row_label)
    for col_idx in range(2, ws.max_column + 1):
        code = _header_code(ws.cell(row=1, column=col_idx).value)
        if code is None:
            continue
        ws.cell(row=target_row, column=col_idx, value=int(holdings_by_code.get(code, 0)))
        ws.cell(row=delta_row, column=col_idx, value=int(delta_by_code.get(code, 0)))
    wb.save(workbook_path)


def _write_detail_sheet(
    workbook_path: Path,
    *,
    sheet_name: str,
    profile_name: str,
    min_weight_floor: float,
    max_weight_cap: float,
    replace_max_weights_with_cap: bool,
    allow_late_start_assets: bool,
    requested_train_start: str,
    requested_train_end: str,
    requested_backtest_start: str,
    requested_backtest_end: str,
    actual_train_start: str,
    actual_backtest_end: str,
    total_value: float,
    optimized_summary: pd.DataFrame,
    current_train: dict[str, float],
    current_test: dict[str, float],
    best: Candidate,
) -> None:
    wb = load_workbook(workbook_path)
    safe_sheet_name = sheet_name[:31]
    for candidate_name in (sheet_name, safe_sheet_name):
        if candidate_name in wb.sheetnames:
            del wb[candidate_name]
    ws = wb.create_sheet(title=safe_sheet_name)

    rows = [
        ["項目", "值"],
        ["profile", profile_name],
        ["min_weight_floor", min_weight_floor],
        ["max_weight_cap", max_weight_cap],
        ["replace_max_weights_with_cap", replace_max_weights_with_cap],
        ["allow_late_start_assets", allow_late_start_assets],
        ["requested_train_range", f"{requested_train_start} ~ {requested_train_end}"],
        ["requested_backtest_range", f"{requested_backtest_start} ~ {requested_backtest_end}"],
        ["actual_train_start", actual_train_start],
        ["actual_backtest_end", actual_backtest_end],
        ["portfolio_market_value", total_value],
        ["optimizer_score", best.score],
        ["optimizer_turnover", best.turnover],
        ["optimizer_beta_vs_0050", best.beta],
        [],
        ["當前投組訓練期", "CAGR", current_train["cagr"], "Sharpe", current_train["sharpe"], "MDD", current_train["max_drawdown"]],
        ["最佳化投組訓練期", "CAGR", best.train_metrics["cagr"], "Sharpe", best.train_metrics["sharpe"], "MDD", best.train_metrics["max_drawdown"]],
        ["當前投組回測期", "CAGR", current_test["cagr"], "Sharpe", current_test["sharpe"], "MDD", current_test["max_drawdown"]],
        ["最佳化投組回測期", "CAGR", best.test_metrics["cagr"], "Sharpe", best.test_metrics["sharpe"], "MDD", best.test_metrics["max_drawdown"]],
        [],
        ["Ticker", "Code", "Name", "Latest Price", "Current Shares", "Current Weight", "Target Weight", "Target Shares", "Delta Shares", "Target Value"],
    ]

    for _, row in optimized_summary.iterrows():
        rows.append(
            [
                row["ticker"],
                row["code"],
                row["name"],
                row["latest_price"],
                row["current_shares"],
                row["current_weight"],
                row["target_weight"],
                row["target_shares"],
                row["delta_shares"],
                row["target_value"],
            ]
        )

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Optimize a full Taiwan holdings workbook.")
    parser.add_argument("--xlsx", required=True, help="Path to the holdings workbook")
    parser.add_argument("--profile", choices=sorted(PROFILE_PRESETS.keys()), default="conservative")
    parser.add_argument("--train-start", default="2020-01-01")
    parser.add_argument("--train-end", default="2024-12-31")
    parser.add_argument("--backtest-start", default="2025-01-01")
    parser.add_argument("--backtest-end", default="2026-05-08")
    parser.add_argument("--download-end", default="2026-05-09")
    parser.add_argument("--samples", type=int, default=400_000)
    parser.add_argument("--seed", type=int, default=7)
    parser.add_argument("--min-beta", type=float, default=None)
    parser.add_argument("--max-beta", type=float, default=None)
    parser.add_argument("--min-weight-floor", type=float, default=0.0, help="minimum target weight applied to each included asset, e.g. 0.01")
    parser.add_argument("--max-weight-cap", type=float, default=1.0, help="maximum target weight applied to each included asset, e.g. 0.60")
    parser.add_argument("--replace-max-weights-with-cap", action="store_true", help="ignore preset per-asset max weights and use max-weight-cap for every included asset")
    parser.add_argument("--allow-late-start-assets", action="store_true", help="use benchmark trading days and treat pre-listing asset returns as cash until they have data")
    parser.add_argument("--require-nonzero-shares", action="store_true", help="enforce at least 1 target share for each included asset")
    parser.add_argument("--exclude-codes", default="", help="comma-separated codes to exclude, e.g. 00631L,00632R")
    parser.add_argument("--sheet-name", default="opt_full_2020_2026")
    parser.add_argument("--target-row-label", default="全投組目標股數")
    parser.add_argument("--delta-row-label", default="全投組建議增減股數")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = (PROJECT_ROOT.parent / xlsx_path).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    _ensure_main_sheet_columns(xlsx_path)

    excluded_codes = {
        code.strip().upper()
        for code in args.exclude_codes.split(",")
        if code.strip()
    }
    invalid_codes = sorted(code for code in excluded_codes if code not in CODE_TO_TICKER)
    if invalid_codes:
        raise ValueError(f"Unsupported exclude codes: {invalid_codes}")

    profile = _resolve_profile(
        args.profile,
        excluded_codes=excluded_codes,
        min_weight_floor=max(0.0, float(args.min_weight_floor)),
        max_weight_cap=min(1.0, float(args.max_weight_cap)),
        replace_max_weights_with_cap=bool(args.replace_max_weights_with_cap),
    )
    min_beta = profile["beta_range"]["min"] if args.min_beta is None else float(args.min_beta)
    max_beta = profile["beta_range"]["max"] if args.max_beta is None else float(args.max_beta)

    holdings, _ = _load_current_holdings(xlsx_path)
    prices = _download_adjusted_close(
        TICKERS,
        args.train_start,
        args.download_end,
        preserve_late_start=bool(args.allow_late_start_assets),
    )
    train_prices = _prepare_price_window(
        prices,
        args.train_start,
        args.train_end,
        allow_late_start_assets=bool(args.allow_late_start_assets),
    )
    test_prices = _prepare_price_window(
        prices,
        args.backtest_start,
        args.backtest_end,
        allow_late_start_assets=bool(args.allow_late_start_assets),
    )
    if len(train_prices) < 252 or len(test_prices) < 60:
        raise RuntimeError("Not enough common train/backtest rows after alignment")

    latest_prices = prices.ffill().iloc[-1]
    holdings["latest_price"] = holdings["ticker"].map(latest_prices)
    holdings["current_value"] = holdings["current_shares"] * holdings["latest_price"]
    total_value = float(holdings["current_value"].sum())
    if total_value <= 0:
        raise RuntimeError("Current holdings market value is zero")
    holdings["current_weight"] = holdings["current_value"] / total_value

    current_weight_series = holdings.set_index("ticker")["current_weight"].reindex(TICKERS).fillna(0.0)
    current_shares_series = holdings.set_index("ticker")["current_shares"].reindex(TICKERS).fillna(0).astype(int)

    train_returns = _returns_from_prices(train_prices, allow_late_start_assets=bool(args.allow_late_start_assets))
    test_returns = _returns_from_prices(test_prices, allow_late_start_assets=bool(args.allow_late_start_assets))
    best = _search_best_candidate(
        train_returns,
        test_returns,
        current_weight_series.to_numpy(dtype=float),
        profile=profile,
        samples=args.samples,
        seed=args.seed,
        min_beta=min_beta,
        max_beta=max_beta,
        allow_late_start_assets=bool(args.allow_late_start_assets),
    )

    target_shares = _allocate_integer_shares(
        total_value,
        best.weights,
        latest_prices.reindex(TICKERS),
        nonzero_tickers={asset["ticker"] for asset in profile["assets"]} if args.require_nonzero_shares else None,
        eligible_tickers={asset["ticker"] for asset in profile["assets"]},
    )
    target_value = target_shares * latest_prices.reindex(TICKERS)
    optimized_summary = pd.DataFrame(
        {
            "ticker": TICKERS,
            "code": [TICKER_TO_CODE[ticker] for ticker in TICKERS],
            "name": [TICKER_TO_NAME[ticker] for ticker in TICKERS],
            "latest_price": latest_prices.reindex(TICKERS).to_numpy(dtype=float),
            "current_shares": current_shares_series.reindex(TICKERS).to_numpy(dtype=int),
            "current_weight": current_weight_series.reindex(TICKERS).to_numpy(dtype=float),
            "target_weight": best.weights,
            "target_shares": target_shares.reindex(TICKERS).to_numpy(dtype=int),
            "delta_shares": (target_shares - current_shares_series.reindex(TICKERS)).to_numpy(dtype=int),
            "target_value": target_value.reindex(TICKERS).to_numpy(dtype=float),
        }
    )

    current_train = _portfolio_metrics(
        train_returns,
        current_weight_series.to_numpy(dtype=float),
        allow_late_start_assets=bool(args.allow_late_start_assets),
    )
    current_test = _portfolio_metrics(
        test_returns,
        current_weight_series.to_numpy(dtype=float),
        allow_late_start_assets=bool(args.allow_late_start_assets),
    )

    _update_main_sheet(
        xlsx_path,
        holdings_by_code=dict(zip(optimized_summary["code"], optimized_summary["target_shares"])),
        delta_by_code=dict(zip(optimized_summary["code"], optimized_summary["delta_shares"])),
        target_row_label=args.target_row_label,
        delta_row_label=args.delta_row_label,
    )
    _write_detail_sheet(
        xlsx_path,
        sheet_name=args.sheet_name,
        profile_name=args.profile,
        min_weight_floor=profile["min_weight_floor"],
        max_weight_cap=profile["max_weight_cap"],
        replace_max_weights_with_cap=profile["replace_max_weights_with_cap"],
        allow_late_start_assets=bool(args.allow_late_start_assets),
        requested_train_start=args.train_start,
        requested_train_end=args.train_end,
        requested_backtest_start=args.backtest_start,
        requested_backtest_end=args.backtest_end,
        actual_train_start=str(train_prices.index.min().date()),
        actual_backtest_end=str(test_prices.index.max().date()),
        total_value=total_value,
        optimized_summary=optimized_summary,
        current_train=current_train,
        current_test=current_test,
        best=best,
    )

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "profile": args.profile,
        "excluded_codes": sorted(excluded_codes),
        "min_weight_floor": profile["min_weight_floor"],
        "max_weight_cap": profile["max_weight_cap"],
        "replace_max_weights_with_cap": profile["replace_max_weights_with_cap"],
        "allow_late_start_assets": bool(args.allow_late_start_assets),
        "require_nonzero_shares": bool(args.require_nonzero_shares),
        "requested_ranges": {
            "train_start": args.train_start,
            "train_end": args.train_end,
            "backtest_start": args.backtest_start,
            "backtest_end": args.backtest_end,
        },
        "actual_ranges": {
            "train_start": str(train_prices.index.min().date()),
            "train_end": str(train_prices.index.max().date()),
            "backtest_start": str(test_prices.index.min().date()),
            "backtest_end": str(test_prices.index.max().date()),
        },
        "portfolio_value": total_value,
        "current_train_metrics": current_train,
        "current_backtest_metrics": current_test,
        "best_candidate": {
            "score": best.score,
            "turnover": best.turnover,
            "beta_vs_0050": best.beta,
            "beta_range": {"min": min_beta, "max": max_beta},
            "bond_range": profile["bond_range"],
            "weights": {ticker: float(weight) for ticker, weight in zip(TICKERS, best.weights)},
            "train_metrics": best.train_metrics,
            "backtest_metrics": best.test_metrics,
        },
        "target_shares": {ticker: int(target_shares[ticker]) for ticker in TICKERS},
    }
    result_path = RESULTS_DIR / (
        f"optimize_excel_holdings_full_{args.profile}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    )
    result_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=" * 72)
    print("Full holdings optimization complete")
    print(f"Workbook: {xlsx_path}")
    print(f"Profile: {args.profile}")
    print(f"Excluded codes: {sorted(excluded_codes)}")
    print(f"Min weight floor: {profile['min_weight_floor']:.2%}")
    print(f"Max weight cap: {profile['max_weight_cap']:.2%}")
    print(f"Replace preset max weights: {profile['replace_max_weights_with_cap']}")
    print(f"Allow late-start assets: {args.allow_late_start_assets}")
    print(f"Require nonzero shares: {args.require_nonzero_shares}")
    print(f"Detail sheet: {args.sheet_name}")
    print(f"Result JSON: {result_path}")
    print(f"Actual train range: {train_prices.index.min().date()} ~ {train_prices.index.max().date()}")
    print(f"Backtest range: {test_prices.index.min().date()} ~ {test_prices.index.max().date()}")
    print(f"Portfolio market value: {total_value:,.2f}")
    print("Target weights:")
    for ticker, weight in zip(TICKERS, best.weights):
        print(f"  {ticker}: {weight:.2%}")
    print("Target shares:")
    for ticker in TICKERS:
        print(f"  {ticker}: {int(target_shares[ticker])}")
    print("Current backtest metrics:")
    print(
        f"  CAGR={current_test['cagr']:.2%} Sharpe={current_test['sharpe']:.3f} "
        f"MDD={current_test['max_drawdown']:.2%}"
    )
    print("Optimized backtest metrics:")
    print(
        f"  CAGR={best.test_metrics['cagr']:.2%} Sharpe={best.test_metrics['sharpe']:.3f} "
        f"MDD={best.test_metrics['max_drawdown']:.2%}"
    )
    print("=" * 72)


if __name__ == "__main__":
    main()
