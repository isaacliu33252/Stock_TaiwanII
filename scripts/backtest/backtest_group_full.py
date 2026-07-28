#!/usr/bin/env python3
"""Classical Max-Sharpe optimizer across the full real 8-ticker holdings ("groupFull").

Research-only. Does not update the latest strategy, live signal, or any
allocation file. Not wired into `execution_plan.py` or the daily pipeline.

Origin: user asked (2026-07-26) to build "groupAB(full)" -> renamed
"groupFull" and pointed at the current real holdings workbook
(`taiwan_stock_20260725.xlsx`, 8 tickers, no more Group A++/Group B split
in the sheet itself). Follows from the arXiv:2601.04062v3 (SPO portfolio
optimization) paper review earlier this session -- see
`GROUP_A_PLUS_2601_04062_SPO_PAPER_REVIEW_HANDOFF_20260726.md` for the full
history of why an optimization layer wasn't judged worth building for
Group A+'s narrow 4-ticker tactical universe, and how "full holdings"
turned out to actually mean 8 real tickers, not 4.

**Design choice, stated explicitly rather than asked** (user rejected a
clarifying-question round and said to just proceed): this v1 uses the
paper's classical Max Sharpe / PtO baseline (Section 3.3.1 of the paper --
`w* = argmax w^T r_bar / sqrt(w^T Sigma w)`, long-only, fully invested,
rolling-window mean/covariance from trailing returns), NOT a trained RL
policy (Group B's old RL policy hasn't run since 2026-06-28 and reviving
it is separate, higher-risk work) and NOT a hand-tuned regime/threshold
system (that's what Group A+/a2118 already is for the tactical sub-slice).
This is the fastest-to-build, most auditable option, and matches the
SPO paper's own PtO baseline exactly -- a reasonable "what would a classical
optimizer recommend for the full real portfolio" reference point. Decision-
focused (SPO+) training was NOT attempted here either: it would require an
end-to-end differentiable optimization layer this project doesn't have
(see the handoff doc) and there is no continuous "predicted return" signal
for these 8 tickers to train one against.

Universe (all 8 tickers currently held per taiwan_stock_20260725.xlsx,
2026-07-25 snapshot): 0050, 0056, 00631L, 00646, 00679B, 00713, 00751B,
00878. 00632R (0 shares in this workbook) is excluded. 00751B is included
in the optimization universe (not hard-excluded) despite the 2026-06-19
finding that it underperforms cash in GroupA++ (see
`GROUP_A_PLUS_PLUS_00751B_CASH_20260619.md`) -- a real optimizer computing
over the full universe should be free to weight it near zero itself if
that finding still holds; hard-excluding it would just be re-asserting the
old conclusion rather than testing it in this new joint context.

Joint price history for all 8 tickers starts 2020-07-10 (00878's IPO,
the latest of the 8) -- see `_yearly_score_ceiling_report`-style caution
from earlier this session
([[feedback_check_data_coverage_before_multiyear_framing]]): this really
is the full usable window, not an arbitrary cutoff, since every other
ticker's data starts earlier.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import duckdb
import numpy as np
import openpyxl
import pandas as pd
from scipy.optimize import minimize

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "FinRL" / "data" / "stock_data.db"
DEFAULT_WORKBOOK = PROJECT_ROOT / "taiwan_stock_20260725.xlsx"

TICKERS = [
    "0050.TW",
    "0056.TW",
    "00631L.TW",
    "00646.TW",
    "00679B.TWO",
    "00713.TW",
    "00751B.TWO",
    "00878.TW",
]
# Workbook column header -> ticker (headers use bare codes, sometimes with a
# typo, e.g. "0063L" instead of "00631L" in the 2026-07-25 sheet).
HEADER_TO_TICKER = {
    "0050": "0050.TW",
    "0056": "0056.TW",
    "00631L": "00631L.TW",
    "0063L": "00631L.TW",
    "00646": "00646.TW",
    "00679B": "00679B.TWO",
    "00713": "00713.TW",
    "00751B": "00751B.TWO",
    "00878": "00878.TW",
}
BOND_ETFS = {"00679B.TWO", "00751B.TWO"}

COMMISSION_RATE = 0.001425
SLIPPAGE_RATE = 0.0005
EQUITY_ETF_SELL_TAX = 0.001


def _load_current_holdings(workbook: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_row = next(r for r in rows if r and any(isinstance(c, str) and "\n" in c for c in r if c))
    holdings_row = next(r for r in rows if r and r[0] == "即時庫存")
    holdings: dict[str, int] = {}
    for header_cell, qty in zip(header_row, holdings_row):
        if not isinstance(header_cell, str):
            continue
        code = header_cell.split("\n")[-1].strip()
        ticker = HEADER_TO_TICKER.get(code)
        if ticker is None:
            continue
        holdings[ticker] = int(qty or 0)
    return holdings


def _load_prices(start: str, end: str) -> pd.DataFrame:
    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        placeholders = ", ".join(["?"] * len(TICKERS))
        rows = con.execute(
            f"""
            SELECT dt, ticker, close
            FROM ohlcv
            WHERE ticker IN ({placeholders}) AND dt BETWEEN ? AND ?
            ORDER BY dt, ticker
            """,
            [*TICKERS, start, end],
        ).fetchdf()
    finally:
        con.close()
    prices = rows.pivot(index="dt", columns="ticker", values="close").sort_index()
    prices.index = pd.to_datetime(prices.index)
    return prices.dropna(subset=TICKERS)


def _max_sharpe_weights(mean_daily_ret: np.ndarray, cov_daily: np.ndarray) -> np.ndarray:
    n = len(mean_daily_ret)

    def neg_sharpe(w: np.ndarray) -> float:
        ret = w @ mean_daily_ret
        vol = np.sqrt(max(w @ cov_daily @ w, 1e-14))
        return -ret / vol

    constraints = [{"type": "eq", "fun": lambda w: w.sum() - 1.0}]
    bounds = [(0.0, 1.0)] * n
    x0 = np.full(n, 1.0 / n)
    result = minimize(neg_sharpe, x0, method="SLSQP", bounds=bounds, constraints=constraints, options={"maxiter": 500, "ftol": 1e-10})
    if not result.success:
        return x0
    w = np.clip(result.x, 0.0, None)
    return w / w.sum()


def _rolling_backtest(
    prices: pd.DataFrame,
    lookback_days: int,
    rebalance_every: int,
    initial_value: float,
) -> dict:
    rets = prices.pct_change().dropna()
    dates = rets.index
    n_assets = len(TICKERS)

    value = initial_value
    weights = np.zeros(n_assets)
    curve = []
    trade_log = []
    total_cost = 0.0
    rebalance_count = 0

    for i in range(lookback_days, len(dates), 1):
        dt = dates[i]
        day_ret = rets.loc[dt, TICKERS].to_numpy()
        value *= 1.0 + float(weights @ day_ret)
        # weights drift with relative performance between rebalances
        if value > 0:
            drifted = weights * (1.0 + day_ret)
            weights = drifted / drifted.sum() if drifted.sum() > 0 else weights

        is_rebalance_day = (i - lookback_days) % rebalance_every == 0
        if is_rebalance_day:
            window = rets.iloc[i - lookback_days : i][TICKERS]
            mean_daily = window.mean().to_numpy()
            cov_daily = window.cov().to_numpy()
            target = _max_sharpe_weights(mean_daily, cov_daily)
            turnover = float(np.abs(target - weights).sum())
            if turnover > 1e-6:
                cost_rate_buy = COMMISSION_RATE
                buy_mask = target > weights
                sell_mask = target < weights
                buy_notional = float(((target - weights)[buy_mask]).sum()) * value
                sell_notional = float(((weights - target)[sell_mask]).sum()) * value
                cost = buy_notional * COMMISSION_RATE + buy_notional * SLIPPAGE_RATE
                for idx, ticker in enumerate(TICKERS):
                    if sell_mask[idx]:
                        sell_tax_rate = 0.0 if ticker in BOND_ETFS else EQUITY_ETF_SELL_TAX
                        notional = (weights[idx] - target[idx]) * value
                        cost += notional * (COMMISSION_RATE + SLIPPAGE_RATE + sell_tax_rate)
                value -= cost
                total_cost += cost
                rebalance_count += 1
                trade_log.append({"date": str(dt.date()), "turnover": turnover, "cost": cost, "weights": dict(zip(TICKERS, target.round(4)))})
            weights = target

        curve.append({"date": str(dt.date()), "value": value})

    curve_df = pd.DataFrame(curve).set_index("date")
    daily_ret = curve_df["value"].pct_change().dropna()
    n_years = len(daily_ret) / 252.0
    annual_return = (curve_df["value"].iloc[-1] / curve_df["value"].iloc[0]) ** (1.0 / n_years) - 1.0 if n_years > 0 else float("nan")
    annual_vol = daily_ret.std() * np.sqrt(252)
    sharpe = annual_return / annual_vol if annual_vol > 0 else float("nan")
    downside = daily_ret[daily_ret < 0]
    sortino = annual_return / (downside.std() * np.sqrt(252)) if len(downside) > 0 and downside.std() > 0 else float("nan")
    running_max = curve_df["value"].cummax()
    drawdown = curve_df["value"] / running_max - 1.0
    max_dd = drawdown.min()

    return {
        "start_date": str(dates[lookback_days].date()),
        "end_date": str(dates[-1].date()),
        "initial_value": initial_value,
        "final_value": float(curve_df["value"].iloc[-1]),
        "annual_return_pct": round(annual_return * 100, 3),
        "annual_vol_pct": round(annual_vol * 100, 3),
        "sharpe": round(float(sharpe), 4),
        "sortino": round(float(sortino), 4),
        "max_drawdown_pct": round(float(max_dd) * 100, 3),
        "rebalance_count": rebalance_count,
        "total_cost": round(total_cost, 2),
        "final_weights": dict(zip(TICKERS, weights.round(4))),
        "trade_log": trade_log,
    }


def _current_recommendation(prices: pd.DataFrame, lookback_days: int, holdings: dict[str, int]) -> dict:
    rets = prices.pct_change().dropna()
    window = rets.iloc[-lookback_days:][TICKERS]
    mean_daily = window.mean().to_numpy()
    cov_daily = window.cov().to_numpy()
    target = _max_sharpe_weights(mean_daily, cov_daily)

    latest_prices = prices.iloc[-1][TICKERS]
    current_value = {t: holdings.get(t, 0) * latest_prices[t] for t in TICKERS}
    total_value = sum(current_value.values())
    current_weights = {t: (current_value[t] / total_value if total_value > 0 else 0.0) for t in TICKERS}

    return {
        "as_of_date": str(prices.index[-1].date()),
        "lookback_days": lookback_days,
        "current_holdings_shares": holdings,
        "current_prices": latest_prices.round(2).to_dict(),
        "current_total_value": round(total_value, 2),
        "current_weights": {t: round(w, 4) for t, w in current_weights.items()},
        "recommended_weights": {t: round(w, 4) for t, w in zip(TICKERS, target)},
        "weight_delta": {t: round(target[i] - current_weights[t], 4) for i, t in enumerate(TICKERS)},
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--start", default="2020-07-10", help="Full joint history start (00878's IPO, the latest of the 8 tickers).")
    parser.add_argument("--end", default=None)
    parser.add_argument("--lookback-days", type=int, default=252, help="Trailing window for mean/covariance estimation.")
    parser.add_argument("--rebalance-every", type=int, default=21, help="Trading days between rebalances (~1 month).")
    parser.add_argument("--initial-value", type=float, default=1_000_000.0)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    end = args.end or datetime.now().strftime("%Y-%m-%d")
    print(f"Loading price history {args.start} -> {end} for {len(TICKERS)} tickers...", file=sys.stderr)
    prices = _load_prices(args.start, end)
    print(f"Joint history: {prices.index[0].date()} -> {prices.index[-1].date()}, {len(prices)} trading days", file=sys.stderr)

    holdings = _load_current_holdings(Path(args.workbook))
    print(f"Loaded holdings from {args.workbook}: {holdings}", file=sys.stderr)

    print("\n=== Rolling Max-Sharpe backtest (groupFull, all 8 tickers) ===")
    backtest = _rolling_backtest(prices, args.lookback_days, args.rebalance_every, args.initial_value)
    for k in ("start_date", "end_date", "final_value", "annual_return_pct", "annual_vol_pct", "sharpe", "sortino", "max_drawdown_pct", "rebalance_count", "total_cost"):
        print(f"  {k}: {backtest[k]}")
    print("  final_weights:")
    for t, w in backtest["final_weights"].items():
        print(f"    {t}: {w:.2%}")

    print("\n=== Current-day recommendation (Max-Sharpe vs. actual holdings) ===")
    rec = _current_recommendation(prices, args.lookback_days, holdings)
    print(f"  as_of_date: {rec['as_of_date']}")
    print(f"  current_total_value: {rec['current_total_value']:,.0f}")
    print(f"  {'ticker':<12}{'current_w':>12}{'recommended_w':>16}{'delta':>10}")
    for t in TICKERS:
        print(f"  {t:<12}{rec['current_weights'][t]:>12.2%}{rec['recommended_weights'][t]:>16.2%}{rec['weight_delta'][t]:>+10.2%}")

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(
            json.dumps({"backtest": backtest, "current_recommendation": rec}, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"\nWrote full report to {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
